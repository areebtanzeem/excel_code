import os
import sys
from os.path import join, dirname, realpath
import pandas
from openpyxl import load_workbook
import datetime
from shutil import copyfile
import numpy as np

if not os.path.exists("files"):
    os.mkdir("files")

path = join(dirname(realpath(__file__)), 'files/')
#########  ALGORITHM   ##########
# First read the file and save it into list
# Second convert the A column into String and datetime object and take diffrences of the consecutive rows
# If difference is greater than 1 then add a new row
# else don't do anythng
##################################

def main(filename):
    df = pandas.read_excel(filename,names=['one', 'two', 'three'])
    writer1 = pandas.ExcelWriter(path+"temp_"+filename, engine='xlsxwriter')
    df.to_excel(writer1, sheet_name='Added_Blank_Space', index=False)
    writer1.save()


    df.drop_duplicates(subset=['one'], keep=False, inplace=True)
    df['difference'] = df.one.diff(1)
    list = df.values.tolist()
    list2 = []
    for i in list:
        if i[3] > datetime.timedelta(seconds=1):
            list2.append(["","",""])
            list2.append([i[0],i[1],i[2]])
        else:
            list2.append([i[0],i[1],i[2]])
    
    df1 = pandas.DataFrame(list2, columns = ['one' , 'two', 'three'])
    writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    df1.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    writer.save()
    print(df)
    #print(df.dtypes)






def add_blank_rows(filename):
    data = []
    data2 = []
    gap_rows = []
    duplicate_index = []
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet1']
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    #for i in range(0,len(data)-1):
    #    diff = data[i+1][0]-data[i][0]
    #    print(data[i+1][0])
    #    print(diff)
    #    data2.append(list(data[i]))
    #    if int(float(str(diff).split(":")[2])) > 1:
    #        data2.append(list(("","","")))
    #    elif int(float(str(diff).split(":")[2])) == 0:
    #        duplicate_index.append(i)
    #        duplicate_index.append(i+1)
    #for line in data2:
        #print(line)
    #    pass
    #print(duplicate_index)
    #for j in duplicate_index:
        #print(j)
    #    del data2[j]   
    #df = pandas.DataFrame(data2,columns=['one', 'two', 'three'])
    #writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    #df.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    #writer.save()

    df2 = pandas.DataFrame(data,columns=['one', 'two', 'three'])
    writer2 = pandas.ExcelWriter(path+"temp_"+filename, engine='xlsxwriter')
    df2.to_excel(writer2, sheet_name='Added_Blank_Space', index=False)
    writer2.save()

def round(a,b):
    c = a-b
    if int(float(str(c).split(":")[-1].split(".")[-1])) > 5:
        return int(float(str(c).split(":")[-1].split(".")[0])) + 1
    else:
        return int(float(str(c).split(":")[-1].split(".")[0]))


def add_blank_rows_two(filename):
    data = []
    data2 = []
    gap_rows = []
    duplicate_index = []
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet1']
    for row in sheet.iter_rows(min_row=2,values_only=True):
        data.append(row)

    for i in range(0,len(data)-1):
        a = data[i+1][0]
        b = data[i][0]
        diff = round(a,b)
        print(data[i][0])
        print(data[i+1][0])
        print(diff,"   time delta   ",datetime.timedelta(seconds=1.5))
        
        data2.append(list(data[i]))
        if diff > 1:
            data2.append(list(("","","")))
        elif diff == 0:
            duplicate_index.append(i)
            duplicate_index.append(i+1)
    for line in data2:
        #print(line)
        pass
    print(duplicate_index)
    for j in duplicate_index:
        del data2[j]   
    df = pandas.DataFrame(data2,columns=['one', 'two','three',"blank","four","five","six"])
    writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    writer.save()






def combine_files(filename1,filename2):

    #workbook1 = load_workbook(filename=filename1)
    #sheet1 = workbook1['Added_Blank_Space']

    #workbook2 = load_workbook(filename=filename2)
    #sheet2 = workbook2['Added_Blank_Space']

    #data1 = []
    #data2 = []
    #final_data = []
    #data1_index = []
    #data2_index = []
    #print("data1 length",len(data1))
    #print("data2 length",len(data2))
    #for row1 in sheet1.iter_rows(values_only=True):
     #   data1.append(list(row1))
    
   # for row2 in sheet2.iter_rows(values_only=True):
   #     data2.append(list(row2))
    
#    for i in range(len(data1)-1):
#        for j in range(len(data2)-1):
#            if data1[i][0] in data2[j]:
#                pass
#            else:
#                data1_index.append(i)
#                data2_index.append(j)


#    print(data2_index)
#    print(data1_index)
    



    # TODO CODE HERE
    # ALGORITHM
    #FIRST COMBINE BOTH
    df1 = pandas.read_excel(filename1)
    df2 = pandas.read_excel(filename2)
    print(df1)
    print(df2)
    s1 = pandas.merge(df1, df2, how='inner', on="one")
    #s1.replace({'NaT': None}).dropna(how="all")
    #s1['one'] = s1['one'].dt.round('S')
    s1.drop_duplicates(subset=['one'], keep=False, inplace=True)
    selected_date = s1[["one"]]
    s1.insert(3,'blank','')
    s1.insert(4,'fourth',selected_date)
    #s1['six'] = s1['one']
    #s1['one'] = pandas.to_datetime(s1['one'])
    writer = pandas.ExcelWriter("merge_sheet1_sheet2.xlsx", engine='xlsxwriter')
    s1.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    print(s1)
    add_blank_rows_two("merge_sheet1_sheet2.xlsx")



def merge_converter(filename):
    df = pandas.read_excel(filename)
    #df.drop_duplicates(subset=['one'], keep=False, inplace=True)
    #df.loc['Total'] = pandas.Series(df.sum())
    df.insert(7,'seven','')
    # PERCENTAGE CALCULATIONS
    df['two_percentage'] = df['two'].apply(lambda a: (a/df['two'].sum())*100 )
    df['three_percentage'] = df['three'].apply(lambda a: (a/df['three'].sum())*100)
    df['five_percentage'] = df['five'].apply(lambda a: (a/df['five'].sum())*100 )
    df['six_percentage'] = df['six'].apply(lambda a: (a/df['six'].sum())*100 )
    
    #PERCENTAGE CALCULATED

    

    #DIFFERENCE 

    df['two_p_diff'] = df.two_percentage.diff()
    df['five_p_diff'] = df.five_percentage.diff()
    
    #DIFFERENCE ENDS

    #SAME SAME

    df['two_p_diff_a'] = df.apply(lambda x: x['two_p_diff'] if x['two_p_diff']*x['five_p_diff'] > 0 else np.NaN, axis=1)
    df['five_p_diff_a'] = df.apply(lambda x: x['five_p_diff'] if x['two_p_diff']*x['five_p_diff'] > 0 else np.NaN, axis=1)
    
    #SaME SAME

    #CONVERTING SAME SAME TO POSITIVE
    df['two_p_diff'] = df['two_p_diff_a']
    df['five_p_diff'] = df['five_p_diff_a']

    del df['five_p_diff_a']
    del df['two_p_diff_a']
    
    

    writer = pandas.ExcelWriter(path+"merge_converter.xlsx", engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    print(df)







filenames = ["sheet1.xlsx","sheet2.xlsx"]
for filename in filenames:
    if not os.path.exists(filename):
        print("Both Sheets names should be sheet1 and sheet2")
        input("Press Enter to exit the program")
        sys.exit()
    else:
        pass
for filename in filenames:
    print(filename)
    #copyfile(filename, path+"temp_"+filename)
    try:

        add_blank_rows(filename)
    except Exception as e:
        print(f"Add Blank Rows error of {filename}")
        print(str(e))
    try:
        main(filename)
    except Exception as e:
        print(f"Main Function error of {filename}")
        print(str(e))
try:
    combine_files(path+"temp_"+filenames[0],path+"temp_"+filenames[1])
except Exception as e:
        print(f"Combine_files error")
        print(str(e))    

if os.path.exists(path+"merge_sheet1_sheet2.xlsx"):
    try:
        merge_converter(path+"merge_sheet1_sheet2.xlsx")
    except Exception as e:
        print(f"Merge Converter error")
        print(str(e))


if os.path.exists(path+"temp_"+filenames[0]):
    os.remove(path+"temp_"+filenames[0])

if os.path.exists(path+"temp_"+filenames[1]):
    os.remove(path+"temp_"+filenames[1])

if os.path.exists("merge_sheet1_sheet2.xlsx"):
    os.remove("merge_sheet1_sheet2.xlsx")
print("*****************************************")
print("*****************************************")

print("                       ")
print("                       ")

print("File converted successfully!!")
print("Press Enter to Continue!")
print("                       ")
print("                       ")

print("*****************************************")
print("*****************************************")
input()
sys.exit()