import os
import sys
from os.path import join, dirname, realpath
import pandas
from openpyxl import load_workbook
import datetime
from shutil import copyfile
import numpy as np
import operator
import math

if not os.path.exists("files"):
    os.mkdir("files")

path = join(dirname(realpath(__file__)), 'files/')
#########  ALGORITHM   ##########
# First read the file and save it into list
# Second convert the A column into String and datetime object and take diffrences of the consecutive rows
# If difference is greater than 1 then add a new row
# else don't do anythng
##################################


def main(filename):
    df = pandas.read_excel(filename, names=['one', 'two', 'three'],header=None)
    writer1 = pandas.ExcelWriter(path+"temp_"+filename, engine='xlsxwriter')
    df.to_excel(writer1, sheet_name='Added_Blank_Space', index=False)
    writer1.save()

    # For B column average and for C column sun
    #print("Duplicate entries for file: ", filename)
    #ids = df["one"]
    #print(pandas.concat(g for _, g in df.groupby("one") if len(g) > 1))

    df2 = df.groupby(['one'], as_index=False).agg(
        {'two': 'mean', 'three': 'sum'})

    # For B column average and for C column sun Ends Here

    #df.drop_duplicates(subset=['one'], keep=False, inplace=True)
    df2['difference'] = df2.one.diff(1)
    list = df2.values.tolist()
    list2 = []
    for i in list:
        if i[3] > datetime.timedelta(seconds=1):
            list2.append(["", "", ""])
            list2.append([i[0], i[1], i[2]])
        else:
            list2.append([i[0], i[1], i[2]])

    df3 = pandas.DataFrame(list2, columns=['one', 'two', 'three'])
    writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    df3.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    writer.save()
    print(df3)
    # print(df.dtypes)


def add_blank_rows(filename):
    data = []
    data2 = []
    gap_rows = []
    duplicate_index = []
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet1']
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # for i in range(0,len(data)-1):
    #    diff = data[i+1][0]-data[i][0]
    #    print(data[i+1][0])
    #    print(diff)
    #    data2.append(list(data[i]))
    #    if int(float(str(diff).split(":")[2])) > 1:
    #        data2.append(list(("","","")))
    #    elif int(float(str(diff).split(":")[2])) == 0:
    #        duplicate_index.append(i)
    #        duplicate_index.append(i+1)
    # for line in data2:
        # print(line)
    #    pass
    # print(duplicate_index)
    # for j in duplicate_index:
        # print(j)
    #    del data2[j]
    #df = pandas.DataFrame(data2,columns=['one', 'two', 'three'])
    #writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    #df.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    # writer.save()

    df2 = pandas.DataFrame(data, columns=['one', 'two', 'three'])
    writer2 = pandas.ExcelWriter(path+"temp_"+filename, engine='xlsxwriter')
    df2.to_excel(writer2, sheet_name='Added_Blank_Space', index=False)
    writer2.save()


def round(a, b):
    c = a-b
    if int(float(str(c).split(":")[-1].split(".")[-1])) > 5:
        return int(float(str(c).split(":")[-1].split(".")[0])) + 1
    else:
        return int(float(str(c).split(":")[-1].split(".")[0]))


def add_blank_rows_two(filename):

    df = pandas.read_excel(filename)
    df['difference'] = df.one.diff(1)

    list = df.values.tolist()
    list2 = []
    for i in list:
        if i[7] > datetime.timedelta(seconds=1):
            list2.append(["", "", "","","","",""])
            list2.append([i[0], i[1], i[2],i[3], i[4], i[5],i[6]])
        else:
            list2.append([i[0], i[1], i[2],i[3], i[4], i[5],i[6]])

    df = pandas.DataFrame(list2, columns=['one', 'two', 'three', "blank", "four", "five", "six"])
    writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    writer.save()


def combine_files(filename1, filename2):


    # TODO CODE HERE
    # ALGORITHM
    # FIRST COMBINE BOTH
    df1 = pandas.read_excel(path+filename1)
    df2 = pandas.read_excel(path+filename2)

    # DROPPING BLANK LINES FROM DF1 AND DF2

    df1.drop_duplicates(subset=['one'], keep=False, inplace=True)
    df2.drop_duplicates(subset=['one'], keep=False, inplace=True)

    print(df1)
    print(df2)
    s1 = pandas.merge(df1, df2, how='inner', on="one")
    #s1.replace({'NaT': None}).dropna(how="all")
    #s1['one'] = s1['one'].dt.round('S')
    s1.drop_duplicates(subset=['one'], keep=False, inplace=True)
    selected_date = s1[["one"]]
    s1.insert(3, 'blank', '')
    s1.insert(4, 'fourth', selected_date)
    #s1['six'] = s1['one']
    #s1['one'] = pandas.to_datetime(s1['one'])
    writer = pandas.ExcelWriter(
        "merge_sheet1_sheet2.xlsx", engine='xlsxwriter')
    s1.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    print(s1)
    add_blank_rows_two("merge_sheet1_sheet2.xlsx")


def merge_converter(filename):
    df = pandas.read_excel(filename)
    #df.drop_duplicates(subset=['one'], keep=False, inplace=True)
    #df.loc['Total'] = pandas.Series(df.sum())
    df.insert(7, 'seven', '')
    # PERCENTAGE CALCULATIONS
    df['two_percentage'] = df['two'].apply(lambda a: (a/df['two'].sum())*100)
    df['five_percentage'] = df['five'].apply(
        lambda a: (a/df['five'].sum())*100)
    df['three_percentage'] = df['three'].apply(
        lambda a: (a/df['three'].sum())*100)
    
    df['six_percentage'] = df['six'].apply(lambda a: (a/df['six'].sum())*100)

    df.insert(12, 'twelve', '')


    # DIFFERENCE

    df['two_p_diff'] = df.two_percentage.diff()
    df['five_p_diff'] = df.five_percentage.diff()
    
    df.insert(15, 'fifteen', '')
    # DIFFERENCE ENDS

    # SAME SAME

    df['two_p_same'] = df.apply(lambda x: x['two_p_diff'] if x['two_p_diff']*x['five_p_diff'] > 0 else np.NaN, axis=1)
    df['five_p_same'] = df.apply(lambda x: x['five_p_diff'] if x['five_p_diff']*x['two_p_diff'] > 0 else np.NaN, axis=1)
    

    #Changing p_diff to p_same and deleting  same

    df['two_p_diff'] = df['two_p_same']
    df['five_p_diff'] = df['five_p_same']
   

    del df['two_p_same']
    del df['five_p_same']
   
    
    ## CHECK IF ALL FOUR COLUMNS HAVE DATA IN IT OTHERWISE ENTER NULL DATA

    #GET VALUE OF ABOVE ROW FOR TWO PERCENTAGE

    index_of_not_null_two_p_diff = df[~df.two_p_diff.isnull()].index.tolist()
    print(index_of_not_null_two_p_diff)

  

    total_rows = df.shape[0] - 3
    blank_list = [np.NaN]*(total_rows)



    #SHOW ROWS TWO THREE FIVE SIX ABOVE AND SAME

    two_p_values = []
    three_p_values = []
    five_p_values = []
    six_p_values = []

    two_p_v_d = []
    three_p_v_d = []
    five_p_v_d = []
    six_p_v_d = []


    for j in index_of_not_null_two_p_diff:
        two_p_values.append('')
        two_p_values.append(df._get_value(j-1, 'two_percentage'))
        two_p_values.append(df._get_value(j, 'two_percentage'))
        two_p_v_d.append(df._get_value(j, 'two_percentage') - df._get_value(j-1, 'two_percentage'))

        three_p_values.append('')
        three_p_values.append(df._get_value(j-1, 'three_percentage'))
        three_p_values.append(df._get_value(j, 'three_percentage'))#
        three_p_v_d.append(df._get_value(j, 'three_percentage'))

        five_p_values.append('')
        five_p_values.append(df._get_value(j-1, 'five_percentage'))
        five_p_values.append(df._get_value(j, 'five_percentage'))#
        five_p_v_d.append(df._get_value(j, 'five_percentage') - df._get_value(j-1, 'five_percentage'))

        six_p_values.append('')
        six_p_values.append(df._get_value(j-1, 'six_percentage'))
        six_p_values.append(df._get_value(j, 'six_percentage'))
        six_p_v_d.append(df._get_value(j, 'six_percentage'))
    
 
    #df['two_p_values'] = two_p_values + another_blank_list*(total_rows_another - len(two_p_values))
    #df['three_p_values'] = three_p_values + another_blank_list*(total_rows_another - len(three_p_values))
    #df['five_p_values'] = five_p_values + another_blank_list*(total_rows_another - len(five_p_values))
    #df['six_p_values'] = six_p_values + another_blank_list*(total_rows_another - len(six_p_values))

    additional1 = pandas.DataFrame({'two_p_values': two_p_values})
    additional2 = pandas.DataFrame({'three_p_values': three_p_values})
    additional3 = pandas.DataFrame({'five_p_values': five_p_values})
    additional4 = pandas.DataFrame({'six_p_values': six_p_values})

    df = pandas.concat([df, additional1], axis=1)
    df = pandas.concat([df, additional2], axis=1)
    df = pandas.concat([df, additional3], axis=1)
    df = pandas.concat([df, additional4], axis=1) 

    df.insert(20, 'twenty', '')

    total_rows_another = df.shape[0]
    another_blank_list = [np.NaN]

    
    #two_p_v_d = sorted(two_p_v_d)
    #three_p_v_d = sorted(three_p_v_d)
    #five_p_v_d = sorted(five_p_v_d)
    #six_p_v_d = sorted(six_p_v_d)

    two_p_v_d_neg = [abs(x) for x in two_p_v_d if x < 0]
    three_p_v_d_neg = []
    
    
    for ind in range(0,len(two_p_v_d_neg)):
        three_p_v_d_neg.append(three_p_v_d[ind])
        #del three_p_v_d[0]
    
    
    five_p_v_d_neg = [abs(x) for x in five_p_v_d if x < 0]
    six_p_v_d_neg = []
    
    for ind in range(0,len(five_p_v_d_neg)):
        six_p_v_d_neg.append(six_p_v_d[ind])
        #del six_p_v_d[0]

    
    two_p_v_d_pos = [x for x in two_p_v_d if x > 0]
    three_p_v_d_pos = []

    for indi in range(len(two_p_v_d_neg),len(two_p_v_d)):
        three_p_v_d_pos.append(three_p_v_d[ind])



    five_p_v_d_pos = [x for x in five_p_v_d if x > 0]
    six_p_v_d_pos = []

    for indi in range(len(five_p_v_d_neg),len(five_p_v_d)):
        six_p_v_d_pos.append(six_p_v_d[ind])

    
    print("####################")
    print(f"length of two_p_v_d {len(two_p_v_d)}")
    print(f"length of two_p_v_d_pos {len(two_p_v_d_pos)}")
    print(f"length of two_p_v_d_neg {len(two_p_v_d_neg)}")
    print("####################")

    print("####################")
    print(f"length of three_p_v_d {len(three_p_v_d)}")
    print(f"length of three_p_v_d_pos {len(three_p_v_d_pos)}")
    print(f"length of three_p_v_d_neg {len(three_p_v_d_neg)}")
    print("####################")

    print("####################")
    print(f"length of five_p_v_d {len(five_p_v_d)}")
    print(f"length of five_p_v_d_pos {len(five_p_v_d_pos)}")
    print(f"length of five_p_v_d_neg {len(five_p_v_d_neg)}")
    print("####################")

    print("####################")
    print(f"length of six_p_v_d {len(six_p_v_d)}")
    print(f"length of six_p_v_d_pos {len(six_p_v_d_pos)}")
    print(f"length of six   _p_v_d_neg {len(six_p_v_d_neg)}")
    print("####################")

    minus_two_three_neg = list(map(operator.sub, two_p_v_d_neg, three_p_v_d_neg))
    minus_two_three_pos = list(map(operator.sub, two_p_v_d_pos, three_p_v_d_pos))

    minus_five_six_neg = list(map(operator.sub, five_p_v_d_neg, six_p_v_d_neg))
    minus_five_six_pos = list(map(operator.sub, five_p_v_d_pos, six_p_v_d_pos))

    df['two_p_v_d'] = two_p_v_d + another_blank_list*(total_rows_another - len(two_p_v_d))
    df['three_p_v_d'] = three_p_v_d + another_blank_list*(total_rows_another - len(three_p_v_d))
    df['five_p_v_d'] = five_p_v_d + another_blank_list*(total_rows_another - len(five_p_v_d))
    df['six_p_v_d'] = six_p_v_d + another_blank_list*(total_rows_another - len(six_p_v_d))

    df.insert(25, 'twenty_fifth', '')

    df['two_p_v_d_abs'] = df['two_p_v_d'].abs()
    df['three_p_v_d_abs'] = df['three_p_v_d'].abs()
    df['five_p_v_d_abs'] = df['five_p_v_d'].abs()
    df['six_p_v_d_abs'] = df['six_p_v_d'].abs()

    df.insert(30, 'thirty', '')

    df['2-3abs'] = df['two_p_v_d_abs'] - df['three_p_v_d_abs']
    df['5-6abs'] = df['five_p_v_d_abs'] - df['six_p_v_d_abs']

    df['2-3abs'] = df['2-3abs'].abs()
    df['5-6abs'] = df['5-6abs'].abs()


    df.insert(33, '33', '')
    two_p_v_d_pos_list = df.apply(lambda x :x['two_p_v_d'] if x['two_p_v_d'] > 0 else np.NaN ,axis = 1).values.tolist()
    two_p_v_d_pos_list = [x for x in two_p_v_d_pos_list if math.isnan(x) == False]
    df['two_p_v_d_pos'] = two_p_v_d_pos_list + another_blank_list*(total_rows_another - len(two_p_v_d_pos_list))

    three_p_v_d_pos_list = df.apply(lambda x :x['three_p_v_d'] if x['two_p_v_d'] > 0 else np.NaN ,axis = 1).values.tolist()
    three_p_v_d_pos_list = [x for x in three_p_v_d_pos_list if math.isnan(x) == False]
    df['three_p_v_d_pos'] = three_p_v_d_pos_list + another_blank_list*(total_rows_another - len(three_p_v_d_pos_list))

    five_p_v_d_pos_list = df.apply(lambda x :x['five_p_v_d'] if x['five_p_v_d'] > 0 else np.NaN ,axis = 1).values.tolist()
    five_p_v_d_pos_list = [x for x in five_p_v_d_pos_list if math.isnan(x) == False]
    df['five_p_v_d_pos'] = five_p_v_d_pos_list + another_blank_list*(total_rows_another - len(five_p_v_d_pos_list))
    
    six_p_v_d_pos_list = df.apply(lambda x :x['six_p_v_d'] if x['five_p_v_d'] > 0 else np.NaN ,axis = 1).values.tolist()
    six_p_v_d_pos_list = [x for x in six_p_v_d_pos_list if math.isnan(x) == False]
    df['six_p_v_d_pos'] = six_p_v_d_pos_list + another_blank_list*(total_rows_another - len(six_p_v_d_pos_list))

    df.insert(38, '38', '')
    two_p_v_d_neg_list = df.apply(lambda x :x['two_p_v_d'] if x['two_p_v_d'] < 0 else np.NaN ,axis = 1).values.tolist()
    two_p_v_d_neg_list = [x for x in two_p_v_d_neg_list if math.isnan(x) == False]
    df['two_p_v_d_neg'] = two_p_v_d_neg_list + another_blank_list*(total_rows_another - len(two_p_v_d_neg_list))
    
    three_p_v_d_neg_list = df.apply(lambda x :x['three_p_v_d'] if x['two_p_v_d'] < 0 else np.NaN ,axis = 1).values.tolist()
    three_p_v_d_neg_list = [x for x in three_p_v_d_neg_list if math.isnan(x) == False]
    df['three_p_v_d_neg'] = three_p_v_d_neg_list + another_blank_list*(total_rows_another - len(three_p_v_d_neg_list))

    five_p_v_d_neg_list = df.apply(lambda x :x['five_p_v_d'] if x['five_p_v_d'] < 0 else np.NaN ,axis = 1).values.tolist()
    five_p_v_d_neg_list = [x for x in five_p_v_d_neg_list if math.isnan(x) == False]
    df['five_p_v_d_neg'] = five_p_v_d_neg_list + another_blank_list*(total_rows_another - len(five_p_v_d_neg_list))

    six_p_v_d_neg_list = df.apply(lambda x :x['six_p_v_d'] if x['five_p_v_d'] < 0 else np.NaN ,axis = 1).values.tolist()
    six_p_v_d_neg_list = [x for x in six_p_v_d_neg_list if math.isnan(x) == False]
    df['six_p_v_d_neg'] = six_p_v_d_neg_list + another_blank_list*(total_rows_another - len(six_p_v_d_neg_list))
    
    df['two_p_v_d_neg'] = df['two_p_v_d_neg'].abs()
    df['three_p_v_d_neg'] = df['three_p_v_d_neg'].abs()
    df['five_p_v_d_neg'] = df['five_p_v_d_neg'].abs()
    df['six_p_v_d_neg'] = df['six_p_v_d_neg'].abs()

    df.insert(43, '43', '')

    df['2-3pos'] = df['two_p_v_d_pos'] - df['three_p_v_d_pos']
    df['5-6pos'] = df['five_p_v_d_pos'] - df['six_p_v_d_pos']

    df['2-3pos'] = df['2-3pos'].abs()
    df['5-6pos'] = df['5-6pos'].abs()
    df.insert(46, '46', '')

    df['2-3neg'] = df['two_p_v_d_neg'] - df['three_p_v_d_neg']
    df['5-6neg'] = df['five_p_v_d_neg'] - df['six_p_v_d_neg']

    df['2-3neg'] = df['2-3neg'].abs()
    df['5-6neg'] = df['5-6neg'].abs()

    #CHOTA WALA RAKHNA HAI BADA WALA HATANA HAI
    df.insert(49, '49', '')

    df['2-3pos_a'] = df.apply(lambda x : x['2-3pos'] if x['2-3pos'] < x['5-6pos'] else np.NaN, axis = 1)
    df['5-6pos_a'] = df.apply(lambda x : x['5-6pos'] if x['5-6pos'] < x['2-3pos'] else np.NaN, axis = 1)

    df['2-3neg_a'] = df.apply(lambda x : x['2-3neg'] if x['2-3neg'] < x['5-6neg'] else np.NaN, axis = 1)
    df['5-6neg_a'] = df.apply(lambda x : x['5-6neg'] if x['5-6neg'] < x['2-3neg'] else np.NaN, axis = 1)

    av_23_pos_a = df['2-3pos_a'].mean()
    av_56_pos_a = df['5-6pos_a'].mean()

    av_23_neg_a = df['2-3neg_a'].mean()
    av_56_neg_a = df['5-6neg_a'].mean()

    av_23_pos_a_percent = av_23_pos_a*100/(av_23_pos_a + av_23_neg_a)
    av_23_neg_a_percent = av_23_neg_a*100/(av_23_pos_a + av_23_neg_a)
    
    av_56_pos_a_percent = av_56_pos_a*100/(av_56_pos_a+av_56_neg_a)
    av_56_neg_a_percent = av_56_neg_a*100/(av_56_pos_a+av_56_neg_a)

    a = df['2-3pos_a'].values.tolist()
    a = [x for x in a if math.isnan(x) == False]
    a.append(av_23_pos_a)
    a.append(av_23_pos_a_percent)


    b = df['5-6pos_a'].values.tolist()
    b = [x for x in b if math.isnan(x) == False]
    b.append(av_56_pos_a)
    b.append(av_56_pos_a_percent)

    c = df['2-3neg_a'].values.tolist()
    c = [x for x in c if math.isnan(x) == False]
    c.append(av_23_neg_a)
    c.append(av_23_neg_a_percent)

    d = df['5-6neg_a'].values.tolist()
    d = [x for x in d if math.isnan(x) == False]
    d.append(av_56_neg_a)
    d.append(av_56_neg_a_percent)
    total_rows_another = df.shape[0]

    print(f'length of a {len(a)}')
    print(f'total rows {total_rows_another}')
    print(f'remainder {total_rows_another - len(a)}')
    df['2-3pos_a'] = a + another_blank_list*(total_rows_another - len(a))
    df['5-6pos_a'] = b + another_blank_list*(total_rows_another - len(b))
    df['2-3neg_a'] = c + another_blank_list*(total_rows_another - len(c))
    df['5-6neg_a'] = d + another_blank_list*(total_rows_another - len(d))



    writer = pandas.ExcelWriter(path+"merge_converter.xlsx", engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    print(df)


filenames = ["sheet1.xlsx", "sheet2.xlsx"]
for filename in filenames:
    if not os.path.exists(filename):
        print("Both Sheets names should be sheet1 and sheet2")
        input("Press Enter to exit the program")
        sys.exit()
    else:
        pass
for filename in filenames:
    print(filename)
    copyfile(filename, path+"temp_"+filename)
    try:
        pass
        # add_blank_rows(filename)
    except Exception as e:
        print(f"Add Blank Rows error of {filename}")
        print(str(e))
    try:
        main(filename)
    except Exception as e:
        print(f"Main Function error of {filename}")
        print(str(e))
try:
    # pass
    combine_files(filenames[0], filenames[1])
except Exception as e:
    print(f"Combine_files error")
    print(str(e))

#if os.path.exists(path+"merge_sheet1_sheet2.xlsx"):
#    try:
merge_converter(path+"merge_sheet1_sheet2.xlsx")
#    except Exception as e:
#        print(f"Merge Converter error")
#        print(str(e))


if os.path.exists(path+"temp_"+filenames[0]):
    os.remove(path+"temp_"+filenames[0])

if os.path.exists(path+"temp_"+filenames[1]):
    os.remove(path+"temp_"+filenames[1])

if os.path.exists("merge_sheet1_sheet2.xlsx"):
    os.remove("merge_sheet1_sheet2.xlsx")
print("*****************************************")
print("*****************************************")

print("                       ")
print("                       ")

print("File converted successfully!!")
print("Press Enter to Continue!")
print("                       ")
print("                       ")

print("*****************************************")
print("*****************************************")
input()
sys.exit()
