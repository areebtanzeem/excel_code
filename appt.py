import os
import sys
from os.path import join, dirname, realpath
import pandas
from openpyxl import load_workbook
import datetime
from shutil import copyfile
import numpy as np
import operator
import math

if not os.path.exists("files"):
    os.mkdir("files")

path = join(dirname(realpath(__file__)), 'files/')
#########  ALGORITHM   ##########
# First read the file and save it into list
# Second convert the A column into String and datetime object and take diffrences of the consecutive rows
# If difference is greater than 1 then add a new row
# else don't do anythng
##################################


def main(filename):
    df = pandas.read_excel(filename, names=['one', 'two', 'three'],header=None)
    writer1 = pandas.ExcelWriter(path+"temp_"+filename, engine='xlsxwriter')
    df.to_excel(writer1, sheet_name='Added_Blank_Space', index=False)
    writer1.save()

    # For B column average and for C column sun
    #print("Duplicate entries for file: ", filename)
    #ids = df["one"]
    #print(pandas.concat(g for _, g in df.groupby("one") if len(g) > 1))

    df2 = df.groupby(['one'], as_index=False).agg(
        {'two': 'mean', 'three': 'sum'})

    # For B column average and for C column sun Ends Here

    #df.drop_duplicates(subset=['one'], keep=False, inplace=True)
    df2['difference'] = df2.one.diff(1)
    list = df2.values.tolist()
    list2 = []
    for i in list:
        if i[3] > datetime.timedelta(seconds=1):
            list2.append(["", "", ""])
            list2.append([i[0], i[1], i[2]])
        else:
            list2.append([i[0], i[1], i[2]])

    df3 = pandas.DataFrame(list2, columns=['one', 'two', 'three'])
    writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    df3.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    writer.save()
    print(df3)
    # print(df.dtypes)


def add_blank_rows(filename):
    data = []
    data2 = []
    gap_rows = []
    duplicate_index = []
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet1']
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # for i in range(0,len(data)-1):
    #    diff = data[i+1][0]-data[i][0]
    #    print(data[i+1][0])
    #    print(diff)
    #    data2.append(list(data[i]))
    #    if int(float(str(diff).split(":")[2])) > 1:
    #        data2.append(list(("","","")))
    #    elif int(float(str(diff).split(":")[2])) == 0:
    #        duplicate_index.append(i)
    #        duplicate_index.append(i+1)
    # for line in data2:
        # print(line)
    #    pass
    # print(duplicate_index)
    # for j in duplicate_index:
        # print(j)
    #    del data2[j]
    #df = pandas.DataFrame(data2,columns=['one', 'two', 'three'])
    #writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    #df.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    # writer.save()

    df2 = pandas.DataFrame(data, columns=['one', 'two', 'three'])
    writer2 = pandas.ExcelWriter(path+"temp_"+filename, engine='xlsxwriter')
    df2.to_excel(writer2, sheet_name='Added_Blank_Space', index=False)
    writer2.save()


def round(a, b):
    c = a-b
    if int(float(str(c).split(":")[-1].split(".")[-1])) > 5:
        return int(float(str(c).split(":")[-1].split(".")[0])) + 1
    else:
        return int(float(str(c).split(":")[-1].split(".")[0]))


def add_blank_rows_two(filename):

    df = pandas.read_excel(filename)
    df['difference'] = df.one.diff(1)
    print("DIFFERENCE COLUMN")
    print(df['difference'])

    list = df.values.tolist()
    list2 = []
    for i in list:
        if i[7] > datetime.timedelta(seconds=1.5):
            list2.append(["", "", "","","","",""])
            list2.append([i[0], i[1], i[2],i[3], i[4], i[5],i[6]])
        else:
            list2.append([i[0], i[1], i[2],i[3], i[4], i[5],i[6]])

    df = pandas.DataFrame(list2, columns=['one', 'two', 'three', "blank", "four", "five", "six"])
    writer = pandas.ExcelWriter(path+filename, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Added_Blank_Space', index=False)
    writer.save()


def combine_files(filename1, filename2):


    # TODO CODE HERE
    # ALGORITHM
    # FIRST COMBINE BOTH
    df1 = pandas.read_excel(path+filename1)
    df2 = pandas.read_excel(path+filename2)

    # DROPPING BLANK LINES FROM DF1 AND DF2

    df1.drop_duplicates(subset=['one'], keep=False, inplace=True)
    df2.drop_duplicates(subset=['one'], keep=False, inplace=True)

    print(df1)
    print(df2)
    s1 = pandas.merge(df1, df2, how='inner', on="one")
    #s1.replace({'NaT': None}).dropna(how="all")
    #s1['one'] = s1['one'].dt.round('S')
    s1.drop_duplicates(subset=['one'], keep=False, inplace=True)
    selected_date = s1[["one"]]
    s1.insert(3, 'blank', '')
    s1.insert(4, 'fourth', selected_date)
    #s1['six'] = s1['one']
    #s1['one'] = pandas.to_datetime(s1['one'])
    writer = pandas.ExcelWriter(
        "merge_sheet1_sheet2.xlsx", engine='xlsxwriter')
    s1.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    print("merged s1")
    print(s1)
    add_blank_rows_two("merge_sheet1_sheet2.xlsx")


    

def combine_files_two(filename1, filename2):


    # TODO CODE HERE
    # ALGORITHM
    # FIRST COMBINE BOTH
    df1 = pandas.read_excel(path+filename1)
    df2 = pandas.read_excel(path+filename2)

    print(df1)
    print(df2)
    s1 = pandas.DataFrame()
    
    s1['one'] = df1['one']
    s1['two'] = df1['two']
    s1['three'] = df1['three']
    
    s1['four'] = df2['one']
    s1['five'] = df2['two']
    s1['six'] = df2['three']
    
    
    s1.insert(3, 'blank', '')
    
    
    writer = pandas.ExcelWriter(
        path+"merge_sheet1_sheet2_two.xlsx", engine='xlsxwriter')
    s1.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    



def merge_converter(filename,count):
    df = pandas.read_excel(filename)
    df.insert(7, 'seven', '')
    # PERCENTAGE CALCULATIONS
    df['i'] = df['one']
    df['j'] = df['two']
    #df['j'] = df['j'].fillna(0)
    df['k'] = df.j.diff()
    #df['j'] = df['j'].replace(0,np.NaN)
    df['k'] = df['k'].replace(0,np.NaN)
    df['l'] = df['three']

    # 

    df['n'] = df['four']
    df['o'] = df['five']
    #df['o'] = df['o'].fillna(0)
    df['p'] = df.o.diff()
    #df['o'] = df['o'].replace(0,np.NaN)
    df['p'] = df['p'].replace(0,np.NaN)
    df['q'] = df['six']
    
    df.insert(12, 'twelve', '')

    df_one = pandas.DataFrame()

    df_one['i'] = df['one']
    df_one['j'] = df['two']
    df_one['k'] = df['k']
    df_one['l'] = df['three']
    df_one['l/k'] = df_one.apply(lambda x : x['l']/x['k'] if x['l'] != 0 and x['k'] != 0 else np.NaN, axis = 1 )
    df_one['l/k_pos'] = df_one.apply(lambda x : x['l/k'] if x['l/k'] > 0 else np.NaN , axis = 1)
    df_one['l/k_neg'] = df_one.apply(lambda x : x['l/k'] if x['l/k'] < 0 else np.NaN , axis = 1)

    df_one['n'] = df['four']
    df_one['o'] = df['five']
    df_one['p'] = df['p']
    df_one['q'] = df['six']
    df_one['q/p'] = df_one.apply(lambda x : x['q']/x['p'] if x['q'] != 0 and x['p'] != 0 else np.NaN, axis = 1 )
    df_one['q/p_pos'] = df_one.apply(lambda x : x['q/p'] if x['q/p'] > 0 else np.NaN , axis = 1)
    df_one['q/p_neg'] = df_one.apply(lambda x : x['q/p'] if x['q/p'] < 0 else np.NaN , axis = 1)


    
    # DELETING THE FIRST ROW
    df_one = df_one.iloc[1:]
    
    #ADDING DF_ONE DATA TO DF AGAIN

    df['s'] = df_one['i']
    df['t'] = df_one['j']
    df['u'] = df_one['k']
    df['v'] = df_one['l']
    df['w'] = df_one['l/k']
    df['x'] = df_one['l/k_pos']
    df['y'] = df_one['l/k_neg']

    df['aa'] = df_one['n']
    df['ab'] = df_one['o']
    df['ac'] = df_one['p']
    df['ad'] = df_one['q']
    df['ae'] = df_one['q/p']
    df['af'] = df_one['q/p_pos']
    df['ag'] = df_one['q/p_neg']

    df.insert(17, '17', '')
    df.insert(25, '25', '')

    df1 = df
    df2 = df
    df3 = df
    df4 = df


    df1 = top_values(df1,5)
    


    df2 = top_values(df2,10)
    #df2.insert(33,'33','')
    #df2.insert(38,'38','')
    #df2.insert(43,'43','')
    #df2.insert(48,'48','')
    #df2.insert(52,'52','')
    
    
    df3 = top_values(df3,20)
    #df3.insert(33,'33','')
    #df3.insert(38,'38','')
    #df3.insert(43,'43','')
    #df3.insert(48,'48','')
    #df3.insert(52,'52','')
    
    df4 = top_values(df4,30)
    #df4.insert(33,'33','')
    #df4.insert(38,'38','')
    #df4.insert(43,'43','')
    #df4.insert(48,'48','')
    #df4.insert(52,'52','')
    
    writer = pandas.ExcelWriter(path+"merge_converter"+str(count)+".xlsx", engine='xlsxwriter')
    df1.to_excel(writer, sheet_name='Sheet1', index=False)
    df2.to_excel(writer, sheet_name='Sheet2', index=False)
    df3.to_excel(writer, sheet_name='Sheet3', index=False)
    df4.to_excel(writer, sheet_name='Sheet4', index=False)
    writer.save()
    #print(df)

def top_values(df,no):
    #print(f"This is dataframe in top values for value {n}")
    #print(df)
    #POSITIVE
    df = df.copy(deep=True)
    x_list = df['x'].values.tolist()
    x_list = [x for x in x_list if math.isnan(x) == False]
    x_list = list(set(x_list))
    x_list = sorted(x_list)
    x_list = x_list[0:no]
    print("X LIST",x_list)
    
    #NEGATIVE LIST
    y_list = df['y'].values.tolist()
    y_list = [x for x in y_list if math.isnan(x) == False]
    y_list = list(set(y_list))
    y_list = sorted(y_list,reverse=True)
    y_list = y_list[0:no]
    print("Y LIST",y_list)
    
    
    #POSITIVE LIST
    af_list = df['af'].values.tolist()
    af_list = [x for x in af_list if math.isnan(x) == False]
    af_list = list(set(af_list))
    af_list = sorted(af_list)
    af_list = af_list[0:no]
    print("AF LIST",af_list)
    
    #NEGATIVE LIST
    ag_list = df['ag'].values.tolist()
    ag_list = [x for x in ag_list if math.isnan(x) == False]
    ag_list = list(set(ag_list))
    ag_list = sorted(ag_list,reverse=True)
    ag_list = ag_list[0:no]
    print("AG LIST",ag_list)
    

    total_rows = df.shape[0]
    blank_list = [np.NaN]

    df['ai'] = x_list + blank_list*(total_rows - len(x_list))
    df['aj'] = y_list + blank_list*(total_rows - len(y_list))
    df['ak'] = af_list + blank_list*(total_rows - len(af_list))
    df['al'] = ag_list + blank_list*(total_rows - len(ag_list))
    
    
    #ABS FOR ABOVE VALUES

    df['an'] = df['ai'].abs()
    df['ao'] = df['aj'].abs()
    df['ap'] = df['ak'].abs()
    df['aq'] = df['al'].abs()
    

    x_sum = []
    y_sum = []
    af_sum = []
    ag_sum = []

    x_sum.append(df['an'].mean())
    y_sum.append(df['ao'].mean())
    af_sum.append(df['ap'].mean())
    ag_sum.append(df['aq'].mean())

    x_sum.append(df['an'].mean()*100/(df['an'].mean()+df['ao'].mean()))
    y_sum.append(df['ao'].mean()*100/(df['an'].mean()+df['ao'].mean()))
    af_sum.append(df['ap'].mean()*100/(df['ap'].mean()+df['aq'].mean()))
    ag_sum.append(df['aq'].mean()*100/(df['ap'].mean()+df['aq'].mean()))
    
    y_sum.append( (df['ao'].mean()*100/(df['an'].mean()+df['ao'].mean())) -  df['an'].mean()*100/(df['an'].mean()+df['ao'].mean()))
    af_sum.append( (df['ap'].mean()*100/(df['ap'].mean()+df['aq'].mean())) - (df['aq'].mean()*100/(df['ap'].mean()+df['aq'].mean())) )

    df['as'] = x_sum + blank_list*(total_rows - len(x_sum))
    df['at'] = y_sum + blank_list*(total_rows - len(y_sum))
    df['au'] = af_sum + blank_list*(total_rows - len(af_sum))
    df['av'] = ag_sum + blank_list*(total_rows - len(ag_sum))

    

    df_one = pandas.DataFrame()
    df_two = pandas.DataFrame()

    df_one['s'] = df['s']
    df_one['t'] = df['t']
    df_one['v'] = df['v']
    df_one['w'] = df['w']

    index_one_pos = []

    for i in x_list:
        index_one_pos.append(df_one[df_one['w']==i].index.tolist())
    #CONVERTING INTO FLAT LIST
    print("INSEX", index_one_pos)
    index_one_pos = [ item for elem in index_one_pos for item in elem]
    
    index_one_neg = []

    for j in y_list:
        index_one_neg.append(df_one[df_one['w']==j].index.tolist())

    print("INSEX", index_one_neg)
    index_one_neg = [ item for elem in index_one_neg for item in elem]

    

    xy_data = []

    for m in index_one_pos:
        xy_data.append(['','',''])
        xy_data.append( [ df_one._get_value(m-1, 's') , df_one._get_value(m-1, 't'), df_one._get_value(m-1, 'v') ] )
        xy_data.append( [ df_one._get_value(m, 's') , df_one._get_value(m, 't'), df_one._get_value(m, 'v') ] )
    
    for n in index_one_neg:
        xy_data.append(['','',''])
        xy_data.append( [ df_one._get_value(n-1, 's') , df_one._get_value(n-1, 't'), df_one._get_value(n-1, 'v') ] )
        xy_data.append( [ df_one._get_value(n, 's') , df_one._get_value(n, 't'), df_one._get_value(n, 'v') ] )

    df_two['aa'] = df['aa']
    df_two['ab'] = df['ab']
    df_two['ad'] = df['ad']
    df_two['ae'] = df['ae']
    
    #SECOND DATA WORK HERE

    index_two_pos = []

    for k in af_list:
        index_two_pos.append(df_two[df_two['ae']==k].index.tolist())

    #CONVERTING INTO FLAT LIST
    print("INSEX", index_two_pos)
    index_two_pos = [ item for elem in index_two_pos for item in elem]

    index_two_neg = []

    for l in ag_list:
        index_two_neg.append(df_two[df_two['ae']==l].index.tolist())
    print("INSEX", index_two_neg)

    index_two_neg = [ item for elem in index_two_neg for item in elem]

    print("INDEXES")
    print(index_two_neg)

    afg_data = []
    
    for m in index_two_pos:
        afg_data.append(['','',''])
        afg_data.append( [ df_two._get_value(m-1, 'aa') , df_two._get_value(m-1, 'ab'), df_two._get_value(m-1, 'ad') ] )
        afg_data.append( [ df_two._get_value(m, 'aa') , df_two._get_value(m, 'ab'), df_two._get_value(m, 'ad') ] )
    
    for n in index_two_neg:
        afg_data.append(['','',''])
        afg_data.append( [ df_two._get_value(n-1, 'aa') , df_two._get_value(n-1, 'ab'), df_two._get_value(n-1, 'ad') ] )
        afg_data.append( [ df_two._get_value(n, 'aa') , df_two._get_value(n, 'ab'), df_two._get_value(n, 'ad') ] )

    df_three = pandas.DataFrame(xy_data,columns=['ax','ay','az'])
    df_four = pandas.DataFrame(afg_data,columns=['bb','bc','bd'])

    df['ax'] = df_three['ax']
    df['ay'] = df_three['ay']
    df['az'] = df_three['az']

    df['bb'] = df_four['bb']
    df['bc'] = df_four['bc']
    df['bd'] = df_four['bd']

    df.insert(33,'33','')
    df.insert(38,'38','')
    df.insert(43,'43','')
    df.insert(48,'48','')
    df.insert(52,'52','')
    
    return df


filenames = ["sheet1.xlsx", "sheet2.xlsx"]
for filename in filenames:
    if not os.path.exists(filename):
        print("Both Sheets names should be sheet1 and sheet2")
        input("Press Enter to exit the program")
        sys.exit()
    else:
        pass
for filename in filenames:
    print(filename)
    copyfile(filename, path+"temp_"+filename)
    try:
        pass
        # add_blank_rows(filename)
    except Exception as e:
        print(f"Add Blank Rows error of {filename}")
        print(str(e))
    try:
        main(filename)
    except Exception as e:
        print(f"Main Function error of {filename}")
        print(str(e))
try:
    #pass
    combine_files(filenames[0], filenames[1])
except Exception as e:
    print(f"Combine_files error")
    print(str(e))

try:
    # pass
    combine_files_two(filenames[0], filenames[1])
except Exception as e:
    print(f"Combine_files error")
    print(str(e))
filename2 = ['merge_sheet1_sheet2_two.xlsx','merge_sheet1_sheet2.xlsx']
count = 1
for naame in filename2:
    try:
        #pass
        merge_converter(path+naame,count)
        count = count + 1        
    except Exception as e:
        print(f"Merge Converter error for filename {naame}")
        print(str(e))
        count = count + 1


if os.path.exists(path+"temp_"+filenames[0]):
    os.remove(path+"temp_"+filenames[0])

if os.path.exists(path+"temp_"+filenames[1]):
    os.remove(path+"temp_"+filenames[1])

if os.path.exists("merge_sheet1_sheet2.xlsx"):
    os.remove("merge_sheet1_sheet2.xlsx")

if os.path.exists(path+"merge_sheet1_sheet2_two.xlsx"):
    os.remove(path+"merge_sheet1_sheet2_two.xlsx")

if os.path.exists(path+"merge_sheet1_sheet2.xlsx"):
    os.remove(path+"merge_sheet1_sheet2.xlsx")


print("*****************************************")
print("*****************************************")

print("                       ")
print("                       ")

print("File converted successfully!!")
print("Press Enter to Continue!")
print("                       ")
print("                       ")

print("*****************************************")
print("*****************************************")
input()
sys.exit()
